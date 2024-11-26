# -*- coding: utf-8 -*-
"""
===============================================================================
PLEASE, READ CAREFULLY!!!!!!!! ================================================
===============================================================================
Include the standard dilution series for your curve in PCR_1 of a batch. On the
CFX manager write them down as 10^0, 10^1, etc. NEVER use 10^ in a sample, this
will screw up this python script.

On each PCR, include a sample mix. This is a random mix of some samples. This 
sample mix is amplified 6 times for each PCR. These samples are used for
normalization within the plates. On the CFX manager, write these samples down 
as 'STD'. Do NOT write anything else or behind it, this will screw up this
python script.

Make sure for follow-up PCRs the threshold is set to the same value as the PCR 
including standard dilution series. Otherwhise, the data is not comparable!

For the export, click in CFX manager on [Export], [Custom Export]. At the top,
select export format as .CSV. Uncheck the box below it named 'Include run 
information header'. For this script you need the Sample Name and Cq, but all
other boxes could be checked if this is desired. Click [Export]. Do this for
all your PCRs. After the analytical run code, make sure there is PCR_#. For all
PCRs the number does not matter, only for the one with the dilution serie, name
this file PCR_1. Save all your .CSVs in one folder. Link to this folder at 
'folder_path'.

In the folder with your .CSVs, also save a .CSV with all your sample names and 
the dilution factor. Name the headers of these columns "Sample" and 'Dilution',
otherwise, this will screw up this python script. Name this file the following:
'analytical_run_code'_dilution_rates.csv. 
===============================================================================
@author: rdebeer
"""
# !!!Variables to set==========================================================
# =============================================================================
folder_path = '//zeus.nioz.nl/mmb/molecular_ecology/mollab_team/Projects/2024/MMB/Nicole/qPCRs/Results/exported_data' # Directory where the CSV files are located
std_curve_copies = 4.06 # What is the copies/µL of the standard *10^x
analytical_run_code = 'NIOZ385&NIOZ386' # Analytical run code generated by DAS
power_to_skip = [0] # Which powers do you not want to use in your standard curve? fill in [1,2] for example if you want to skip 10^1 & 10^2
threshold_cq = 5.0 
threshold_copies = 40.0
# =============================================================================
# Import statements ===========================================================
# =============================================================================
import pandas as pd, numpy as np
import glob
import os
import statistics
import math
from scipy import stats
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, PatternFill, Alignment, Border, Side
# =============================================================================
# Creating output =============================================================
# =============================================================================    
# Creates a new folder where are the results will be placed and checks if this
# already exists
newpath = folder_path + r'/results' 
if not os.path.exists(newpath):
    os.makedirs(newpath)
# =============================================================================
# Loading in the raw data =====================================================
# =============================================================================
# List of file names
# Only keeps the file names if the name contains 'PCR_'
file_names = glob.glob(os.path.join(folder_path, analytical_run_code + '_*.csv'))
file_names = [file for file in file_names if 'PCR_' in os.path.basename(file)]

# Empty dictionary for the DataFrames
PCRs = {}

# Load DataFrames and save them with the name 'PCR' followed by the number from the file name
for file in file_names:
    # Extract the number from the file name
    df_name = os.path.basename(file).split('_')[-1].split('.')[0]
    # Read the CSV file and select only 'Sample' and 'Cq' columns
    PCR = pd.read_csv(file, usecols=['Sample', 'Cq'])
    # Save the DataFrame with the name 'PCR' followed by the number
    PCRs[f'PCR_{df_name}'] = PCR 
# =============================================================================
# Normalize data ==============================================================
# =============================================================================
# Creates an empty dataframe 
STDs_df = pd.DataFrame()
# Loops through the dataframs in the dict to extract the STD samples and do the
# correction. For PCR1, the correction is 0. At the end, the STD samples will
# be added to a new final dataframe which will be put into the final file.
for PCR_name, PCR_df in PCRs.items():
    if PCR_name == 'PCR_1':
        STDs_from_PCR1_df = PCRs['PCR_1'][PCRs['PCR_1']['Sample'] == 'STD']
        average_STD_PCR1 = statistics.mean(STDs_from_PCR1_df['Cq'])
        STD_correction_PCR1 = 0
        PCR_df['Corrected_Cq'] = PCR_df['Cq'] - STD_correction_PCR1
        STDs = STDs_from_PCR1_df["Cq"].values.tolist()
        STDs_df[PCR_name] = pd.Series(STDs) # pd.Series is used for different lenghts of lists.
        
    elif PCR_name != 'PCR_1':
        # Filter STD samples from the current DataFrame
        STDs_from_df = PCR_df[PCR_df['Sample'] == 'STD']
        # Calculate correction based on the average Cq value of PCR_1
        STD_correction = statistics.mean(STDs_from_df['Cq']) - average_STD_PCR1
        # Apply correction to the Cq values of the current DataFrame
        PCR_df['Corrected_Cq'] = PCR_df['Cq'] - STD_correction
        STDs = STDs_from_df["Cq"].values.tolist()
        STDs_df[PCR_name] =  pd.Series(STDs) # pd.Series is used for different lenghts of lists.

# Rename columns to Cq# (e.g., Cq1, Cq2, ...)
STDs_df.index = [f'Cq{i+1}' for i in range(STDs_df.shape[0])]

# .transpose() flips the columns and rows
STD_df_final = STDs_df.transpose()
# Calculates the mean per row and added this in the final column
STD_df_final["Mean"] = STD_df_final.mean(axis=1)
# Calculates the differences between the mean.
STD_df_final["Correction"] = STD_df_final["Mean"] - average_STD_PCR1
# =============================================================================
# Making a standard curve =====================================================
# =============================================================================
# Select the DataFrame 'stdcurve' by filtering on 'Sample' starting with "10^"
stdcurve = PCRs['PCR_1'][PCRs['PCR_1']['Sample'].str.startswith("10^")]
# Remove the rows from 'PCR_1' that are already used in 'stdcurve'
PCR_1 = PCRs['PCR_1'].drop(stdcurve.index)
# Update the 'PCR_1' DataFrame in the PCRs dictionary
PCRs['PCR_1'] = PCR_1

stdcurve['Power'] = (stdcurve["Sample"].str.split('^', expand=True)[1].astype(float))

# For each standard, calculate + add copies/µL and log_copies to the dataframe
for standard in stdcurve.index:
    power = stdcurve['Power'][standard]
    copies = std_curve_copies * 10 ** power
    log_copies = math.log10(copies)
    stdcurve.loc[standard, 'copies/µL'] = copies
    stdcurve.loc[standard, 'log_copies'] = log_copies
    
skipped_values = pd.DataFrame()
stdcurve_calculations = stdcurve
for power in power_to_skip:
    stdcurve_calculations = stdcurve_calculations.drop(stdcurve_calculations[stdcurve_calculations['Power'] == power].index)
    new_row = stdcurve.loc[stdcurve['Power'] == power]
    skipped_values = pd.concat([skipped_values, new_row])

# Drops all the reactions where the Cq is Nan
stdcurve_calculations = stdcurve_calculations.dropna()

# Linear regression + interpolation for standard curve
slope, yintercept, r2, pv, se = stats.linregress(stdcurve_calculations["log_copies"], 
                                                  stdcurve_calculations["Cq"])
interp = np.linspace(np.min(stdcurve_calculations['log_copies']), 
                            np.max(stdcurve_calculations['log_copies']), 
                            num=500)
# calculate efficiency
efficiency = (-1+10**(-1/slope))*100

# determine highest standard
max_power = max(stdcurve_calculations['Power'])

# plot standard curve
fig, ax = plt.subplots(dpi=300)                 # empty plot
ax.set_xticks(np.arange(0,max_power + 1,1))     # x-grid intervals=1
ax.grid(alpha=0.3)                              # transparancy of grid
ax.set_title('Standard curve', fontsize=16)

# Scatterplot for the points included in the analysis
ax.scatter(stdcurve_calculations["log_copies"], # x-axis
            stdcurve_calculations["Cq"],        # y-axis
            c='lime',                           # color of the dots
            label = "Included in analysis")     # labl for legend

# Scatterplot for the points excluded in the analysis
if power_to_skip:
    ax.scatter(skipped_values["log_copies"],    # x-axis
            skipped_values["Cq"],               # y-axis
            c='red',                            # color of the dots
            label = "Excluded in analysis")     # labl for legend                         

# Set labels of the axis
ax.set_xlabel('log10 copies')                   # x-axis label
ax.set_ylabel('Cq')                             # y-axis label

# plot linear regression
ax.plot(interp,                     # x-axis
        yintercept + slope * interp,# y-axis
        linestyle='--',             # style of the line
        c='cornflowerblue'          # color of the line
        )

# add equation to plot
equation = "y = " + str(round(slope, 3)) + "X + " + str(round(yintercept,2))
ax.text(.7, 0.9, equation, 
        size=8, color='purple', 
        transform=ax.transAxes
        )
# add efficiency to plot
efficiency = "efficiency = " + str(round(efficiency)) + "%"
ax.text(.7, 0.85, efficiency, 
        size=8, color='purple', 
        transform=ax.transAxes)

# add r2 to plot
r = "R" + "\u00b2" + f"= {r2:.4f}"  # Format with 2 decimal places
ax.text(.7, 0.8, r, 
        size=8, color='purple', 
        transform=ax.transAxes)

# Adds a legend on the lower left corner
plt.legend(loc='lower left')

# make layout fit better and save as a .png
plt.tight_layout()
plt.savefig(newpath +"/" + analytical_run_code + "_standard_curve.png")
# =============================================================================
# Sample calculations =========================================================
# =============================================================================
# Creates an empty Dataframe with the columns Sample and Corrected_Cq
combined_data = pd.DataFrame(columns=['Sample', 'Corrected_Cq'])

# Iterates over each DataFrame in PCRs and adds every line that is not STD or
# starts with 10^ and adds them to the empty DataFrame above
for PCR_name, PCR_df in PCRs.items():
    for index, row in PCR_df.iterrows():
        if row['Sample'] != 'STD' or row['Sample'].startswith('10^'):
            combined_data.loc[len(combined_data)] = row[['Sample', 'Corrected_Cq']]

# Combine duplicate measurements
sample_calculations = combined_data.groupby('Sample')['Corrected_Cq'].apply(list).reset_index()

test = sample_calculations
for index, row in sample_calculations.iterrows():
    # Iterate over each element in the list
    Cqs = []
    for i, value in enumerate(row['Corrected_Cq']):
        # Create a new column if it doesn't exist
        col_name_cq = f'Corrected_Cq_{i + 1}'
        if col_name_cq not in sample_calculations.columns:
            sample_calculations[col_name_cq] = None  # or np.nan
        # Assign the value to the new column
        sample_calculations.at[index, col_name_cq] = value
        Cqs.append(value)
    sample_calculations.at[index, 'Mean_Cq'] = mean_cq = statistics.mean(Cqs)
    sample_calculations.at[index, 'Stdev_Cq'] = std_cq = np.std(Cqs, ddof=1)
    sample_calculations.at[index, "CV%_Cq"] = (std_cq/mean_cq)*100
    
    copies = []
    for cq in Cqs: 
        copie = 10**((cq - yintercept) / slope)
        copies.append(copie)
    
    std_copies =  np.std(copies)
    mean_copies = statistics.mean(copies)
    sample_calculations.at[index, 'CV%_copies'] = (std_copies/mean_copies)*100

# Drop the original 'Corrected_Cq' column.
sample_calculations.drop(columns=['Corrected_Cq'], inplace=True)

# Changes the order of the dataframe
all_headers = sample_calculations.columns
corrected_Cq_headers = []
other_headers = []
for header in all_headers:
    if header.startswith('Corrected_Cq'):
        corrected_Cq_headers.append(header)
    else:
        other_headers.append(header)
sample_calculations = sample_calculations[corrected_Cq_headers + other_headers]

# Creates the pathway to the dilution file
dilution_file = folder_path +'/'+ analytical_run_code +'_dilution_rates.csv'

dilution_file_check = os.path.exists(dilution_file)

if dilution_file_check:
    # Reads the dilution rates file and uses the columns named 'Sample' and 'Dilution'
    dilutions = pd.read_csv((dilution_file), usecols=['Sample', 'Dilution'])
   
    # Merges the 2 dataframes and overwrites the sample_calculations column
    sample_calculations = pd.merge(dilutions, sample_calculations, on = 'Sample', how='left')
    
else:    
    print('You did not upload a dilution file. Therefor, all the sample are calculated as undiluted samples! If you did dilute your samples, please upload a dilution file in the folder where your raw data is located.')
    sample_calculations['Dilution'] = '1'

for sample in sample_calculations.index:
    # Calculate from std curve formula (10** because using log-copies)
    copies = 10**((sample_calculations["Mean_Cq"][sample] - yintercept) / slope)
    try:
    # Multiply by dilution factor
        copies = copies * sample_calculations["Dilution"][sample]
    except:
        pass
    # Add to dataframe, use scientific format, 2 decimal points
    sample_calculations.loc[sample, "Extract_copies/µL"] = (
        "{:.2e}".format(copies))
# =============================================================================
# Creating CV% plots ==========================================================
# =============================================================================
# Creates a list with all the sample names
sample_names_list = sample_calculations['Sample'].tolist()
# Creates the plot
fig, ax3 = plt.subplots(dpi=300,figsize = (len(sample_names_list)/6,len(sample_names_list)/15))
ax4 = ax3.twinx()

# Sets information for the 2 plots
ax3.scatter(x=sample_names_list, y= sample_calculations['CV%_Cq'], color= "r", label = 'CV% Cq')
ax4.scatter(x=sample_names_list, y= sample_calculations['CV%_copies'], color= "b", label = 'CV% copies')

# Specify x-axis and y-axis labels 
ax3.set_xticklabels(sample_names_list, rotation=90)  # Set x-axis labels with rotation
ax3.set_ylabel('CV% Cq', fontsize = 20)  # Set y-axis label
ax3.tick_params(axis="y", labelcolor = "r", labelsize = 20)

# Sets label for the y-axis for the copies CV% 
ax4.set_ylabel('CV% copies', fontsize = 20)
ax4.tick_params(axis="y", labelcolor = "b", labelsize = 20)

# Adding the treshold lines
ax3.axhline(threshold_cq, label = "Threshold Cq", color = 'r', linestyle = '--')
ax4.axhline(threshold_copies, label = "Threshold copies", color = 'b', linestyle = '--')

# Creates the title and the legend
fig.suptitle('CV% of the Cq and copies/µL extract added per sample', fontsize = 40)  # Set title for the plot
fig.legend(loc="upper right", bbox_to_anchor=(1,1), bbox_transform=ax3.transAxes, fontsize = 20)

# Makes the plot nicer, saves it and shows it
plt.tight_layout()
plt.savefig(newpath +"/" + analytical_run_code + "_CV_percentages.png")
plt.show()
# =============================================================================
# Generating output ===========================================================
# =============================================================================
# Preparing the dataframes as prefered
sample_calculations.set_index("Sample", inplace = True)

stdcurve.set_index("Sample", inplace = True)
for PCR_name, PCR_df in PCRs.items():
    PCR_df.set_index("Sample", inplace = True)

# Creates a new excel file
excel_file_path = newpath + "/" + analytical_run_code + "_results.xlsx"

# Add the dataframe with the standard curve to the excel file
with pd.ExcelWriter(excel_file_path) as writer:
    sample_calculations.to_excel(writer, sheet_name = "Final_data")
    stdcurve.to_excel(writer, sheet_name = "Standard_curve")
    STD_df_final.to_excel(writer, sheet_name = "STD_samples")    
    for PCR_name, PCR_df in PCRs.items():
        PCR_df.to_excel(writer, sheet_name = f"Raw_data_{PCR_name}")
# =============================================================================
# Formating output ============================================================
# =============================================================================
wb = load_workbook(excel_file_path)
wb2 = load_workbook(folder_path + "/" + analytical_run_code + "_project_info.xlsx")

# Creating a new sheet for the project_info
ws = wb.create_sheet(title = "Project_info")
ws2 = wb2.worksheets[0]

for row in ws2:
    for cell in row:
        ws[cell.coordinate].value = cell.value
max_length = 0
for cell in ws["A"]:
    cell.font = Font(bold=True)
    try:
        if len(str(cell.value)) > max_length:
            max_length = len(cell.value)
    except:
        pass
adjusted_width = (max_length + 2)
ws.column_dimensions["A"].width = adjusted_width

# Changes in the worksheet Final_data
ws = wb['Final_data']
excluded_columns = ['Sample', 'Dilution', 'Extract_copies/µL']

# Get the headers (assumes headers are in the first row)
headers = [cell.value for cell in ws[1]]

# Iterate over columns by index and header
for idx, header in enumerate(headers, start=1):  # Start indexing from 1 for Excel column reference
    if header not in excluded_columns:
        for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):  # Skip the header row
            for c in cell:
                c.number_format = '#,##0.00'  # Format to two decimal places

# Creates a color that is used through out the rest of this protocol.
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')
left_border = Border(left=Side(style = 'thin'))
columns_for_layout = ["Sample", "Corrected_Cq_1", "Mean_Cq", "CV%_Cq", "Extract_copies/µL"]
          
# Gives the cells with a % higher than the thresholds a color
for column in ws.iter_cols():
    if column[0].value == "CV%_Cq":
        for cell in column[1:]:
            if cell.value is not None and cell.value > threshold_cq:
                cell.fill = redFill
    if column[0].value == "CV%_copies":
        for cell in column[1:]:
            if cell.value is not None and cell.value > threshold_copies:
                cell.fill = redFill
    # Gives the columns with the header name in columns_for_layout a thin left border
    if column[0].value in columns_for_layout:
        for cell in column[1:]:
            cell.border = left_border
    if column[0].value == "Extract_copies/µL":
        for cell in column[1:]:
            if cell.value == "nan":
                cell.value = ""
             
# Set a fixed width for all columns with data
fixed_width = 15  # Fixed width to apply
for col in range(1, ws.max_column + 1):  # Iterate over all column indices
    col_letter = get_column_letter(col)  # Convert index to column letter
    ws.column_dimensions[col_letter].width = fixed_width  # Set the column width

# Adding new sheet and adding the chart in the sheet
ws = wb.create_sheet(title = "CV%_graph")
png2 = Image(newpath +"/" + analytical_run_code + "_CV_percentages.png")
png2.height = 440
png2.width = 1160
ws.add_image(png2, "A1")

# Changes in the worksheet Standard_curve
ws = wb['Standard_curve']

# Adding the standard curve into the excel
png = Image(newpath +"/" + analytical_run_code + "_standard_curve.png")

# Changing the height and width of the png
png.height = 500
png.width = 666
ws.add_image(png, "H2")

for column in ws.iter_cols():
    if column[0].value == "Power":
        for cell in column[1:]:
            if cell.value in power_to_skip:
                # Apply red fill to the entire row
                for row_cell in ws[cell.row]:  # Loop through all cells in the same row
                    row_cell.fill = redFill  # Apply red fill to the entire row

# Ordering the workingsheets
mandatory_sheets = ["Project_info", "Final_data", "Standard_curve", "STD_samples", "CV%_graph"]
other_sheets = [sheetname for sheetname in wb.sheetnames if sheetname not in mandatory_sheets]
all_sheets = mandatory_sheets + other_sheets

# Reordering sheets
wb._sheets = [wb[sheet] for sheet in all_sheets]

# Save the workbook with reordered sheets
wb.save(excel_file_path)